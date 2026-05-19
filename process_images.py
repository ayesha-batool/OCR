import os
import re
import tempfile
import threading
import time
import asyncio
from base64 import b64encode
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, HttpUrl
from fastapi.middleware.cors import CORSMiddleware
load_dotenv()

# Folder containing images
BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"

# Supported image formats
SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".webp"}
PDF_EXTENSIONS = {".pdf"}
WORD_EXTENSIONS = {".docx"}
LEGACY_WORD_EXTENSIONS = {".doc"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}

# Matches page headers written by extract_from_pdf (one block per PDF page).
_PAGE_HEADER_RE = re.compile(r"^--- Page (\d+) ---\s*$", re.MULTILINE)

_OCR_PROMPT = (
    "Strict OCR only. Copy every character visible in the image exactly as printed.\n"
    "Output ONLY the raw text from the image in reading order (Arabic, Urdu, English, numbers).\n"
    "Preserve line breaks, punctuation, spelling mistakes, and repeated words exactly.\n"
    "Do NOT translate, summarize, paraphrase, or fix spelling/grammar.\n"
    "FORBIDDEN in your reply — do not output any of these:\n"
    "- Role, Task, Constraints, Block, or descriptions of the image or layout\n"
    "- Labels you invent such as Arabic:, Urdu:, Block 1, or bullet lists of rules\n"
    "- Commentary, reasoning, double-checking, or repeating these instructions\n"
    "- Markdown formatting, asterisks, or meta text about what you are doing\n"
    "If a word like Arabic or Urdu is not printed in the image, do not write it.\n"
    "Reply with extracted text only. No other words before or after.\n"
)

_OCR_META_LINE_RE = re.compile(
    r"^\s*(\*+\s*)?"
    r"(Role\s*:|Task\s*:|Constraints?\s*:|Block\s+\d+\s*:|"
    r"The image contains|No translation|No fixing|Output ONLY|"
    r"Preserve original|Keep Urdu|Do NOT translate|Do not output|"
    r"FORBIDDEN|REQUIRED|Strict OCR|Extract text VERBATIM)"
    r".*$",
    re.IGNORECASE,
)
_OCR_LABEL_PREFIX_RE = re.compile(r"^\s*(Arabic|Urdu)\s*:\s*", re.IGNORECASE)


def _sanitize_ocr_output(text: str) -> str:
    """Drop Gemini meta-commentary; keep verbatim lines from the page."""
    if not text or not text.strip():
        return text

    cleaned: list[str] = []
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            cleaned.append("")
            continue
        if _OCR_META_LINE_RE.match(stripped):
            continue
        if _OCR_LABEL_PREFIX_RE.match(stripped):
            stripped = _OCR_LABEL_PREFIX_RE.sub("", stripped).strip()
            if not stripped:
                continue
        cleaned.append(line)

    result = "\n".join(cleaned).strip()
    while "\n\n\n" in result:
        result = result.replace("\n\n\n", "\n\n")
    return result


# Gemini (used for image extraction)
FREE_MODELS = [
# 200
"gemini-flash-latest",
"gemini-2.5-flash",
"gemini-robotics-er-1.6-preview",
"gemma-3-27b-it",
"gemma-4-26b-a4b-it",
"gemma-4-31b-it",
"gemini-2.5-flash-lite",
"gemma-3-4b-it",
"gemma-3-12b-it",
"gemma-3-27b-it",
"gemma-4-26b-a4b-it",
"gemma-4-31b-it",
"gemini-3-flash-preview",

# 503
"gemini-flash-lite-latest",
"gemini-3.1-flash-lite-preview",

# 429
"gemini-2.0-flash-lite",
"gemini-2.5-pro-preview-tts",
"gemini-pro-latest",
"gemini-2.5-flash-image",
"gemini-2.5-pro",
"gemini-2.0-flash",
"gemini-2.0-flash-001",
"gemini-2.0-flash-lite-001",
"gemini-2.5-computer-use-preview-10-2025",
"gemini-3-pro-preview",
"gemini-3.1-pro-preview",
"gemini-3.1-pro-preview-customtools",
"gemini-3.1-flash-image-preview",
"nano-banana-pro-preview",
"gemini-3-pro-image-preview",
"lyria-3-clip-preview",
"lyria-3-pro-preview",


   
]
API_KEYS = [os.getenv(f"GEMINI_API_KEY_{i}") for i in range(1, 12)]
API_KEYS = [k for k in API_KEYS if k]
PARALLEL_WORKERS = max(1, min(8, int(os.getenv("PARALLEL_WORKERS", "4"))))
_LOG_LOCK = threading.Lock()
_FILE_WRITE_LOCK = threading.Lock()
PROCESSOR_POST_ENDPOINT = os.getenv("PROCESSOR_POST_ENDPOINT", "").strip()
PROCESSOR_GET_ENDPOINT = os.getenv("PROCESSOR_GET_ENDPOINT", "").strip()


_DEFAULT_CORS_ORIGINS = (
    "http://localhost:4200,"
    "https://www.steamx.pk,"
    "https://steamx.pk,"
    "https://steamx-v1-backend.onrender.com,"
    "https://ocr-9e8w.onrender.com"
)


def _cors_origins(default: str = _DEFAULT_CORS_ORIGINS) -> list[str]:
    raw = os.getenv("CORS_ORIGINS", default).strip()
    if raw == "*":
        return ["*"]
    return [origin.strip().rstrip("/") for origin in raw.split(",") if origin.strip()]


app = FastAPI(title="Universal Text Extractor", version="1.0.0")
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins(),
    allow_methods=["*"],
    allow_headers=["*"],
)

class SendExtractedPayload(BaseModel):
    endpoint: HttpUrl
    text: str


class ProcessExtractedPayload(BaseModel):
    text: str


def _log(message: str) -> None:
    with _LOG_LOCK:
        print(message, flush=True)


def _api_key_label(api_key: str | None) -> str:
    if not api_key:
        return "none"
    try:
        slot = API_KEYS.index(api_key) + 1
        return f"GEMINI_API_KEY_{slot}"
    except ValueError:
        return "custom"


def _extract_with_gemini(
    path: Path,
    *,
    preferred_api_key: str | None = None,
    worker_id: int | None = None,
    page_label: str = "",
) -> str:
    worker_tag = f"[Worker {worker_id}] " if worker_id is not None else ""
    page_tag = f"{page_label} " if page_label else ""
    _log(f"\n{worker_tag}📸 {page_tag}OCR {path.name}")

    if not API_KEYS:
        raise RuntimeError("Gemini API keys not configured. Set GEMINI_API_KEY_1..GEMINI_API_KEY_7.")
    if not FREE_MODELS:
        raise RuntimeError("No Gemini models configured.")

    keys_to_try: list[str] = []
    if preferred_api_key and preferred_api_key in API_KEYS:
        keys_to_try.append(preferred_api_key)
    keys_to_try.extend(k for k in API_KEYS if k not in keys_to_try)

    image_b64 = b64encode(path.read_bytes()).decode("utf-8")

    mime = "image/jpeg"
    if path.suffix.lower() == ".png":
        mime = "image/png"

    body = {
        "contents": [
            {
                "role": "user",
                "parts": [
                    {"text": _OCR_PROMPT},
                    {"inline_data": {"mime_type": mime, "data": image_b64}},
                ],
            }
        ],
        "generationConfig": {
            "temperature": 0,
            "topP": 0,
            "topK": 1,
            "responseMimeType": "text/plain",
            "maxOutputTokens": 8192
        },
    }

    last_error: Exception | None = None

    for api_key in keys_to_try:
        key_label = _api_key_label(api_key)
        for model in FREE_MODELS:
            _log(f"{worker_tag}🔑 {key_label} | 🤖 {model}")
            try:
                r = requests.post(
                    f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent",
                    headers={"x-goog-api-key": api_key},
                    json=body,
                    timeout=120,
                )
                _log(f"{worker_tag}🌐 {key_label} | Status: {r.status_code}")
                if r.status_code == 429:
                    _log(f"{worker_tag}🚫 {key_label} rate limit -> next key/model")
                    continue
                r.raise_for_status()
                data = r.json()
                text = (
                    data.get("candidates", [{}])[0]
                    .get("content", {})
                    .get("parts", [{}])[0]
                    .get("text", "")
                    .strip()
                )
                if text:
                    text = _sanitize_ocr_output(text)
                    if not text:
                        _log(f"{worker_tag}⚠️ Empty after removing OCR meta from {key_label}")
                        continue
                    preview = text[:120].replace("\n", " ")
                    _log(f"{worker_tag}✅ {page_tag}OCR done ({len(text)} chars): {preview}…")
                    return text
                _log(f"{worker_tag}⚠️ Empty response from {key_label}")
            except Exception as exc:
                _log(f"{worker_tag}❌ {key_label} exception: {exc}")
                last_error = exc
                continue

    if last_error:
        _log(f"{worker_tag}❌ OCR failed for {path.name}")
        raise RuntimeError(f"Gemini extraction failed across all model/API combinations: {last_error}") from last_error
    _log(f"{worker_tag}❌ OCR failed for {path.name}")
    raise RuntimeError("Gemini extraction failed across all model/API combinations.")


def process_image(
    path: Path,
    *,
    preferred_api_key: str | None = None,
    worker_id: int | None = None,
    page_label: str = "",
) -> str:
    text = _extract_with_gemini(
        path,
        preferred_api_key=preferred_api_key,
        worker_id=worker_id,
        page_label=page_label,
    )
    return text.strip()


def _process_pdf_page(
    file_bytes: bytes,
    page_number: int,
    total_pages: int,
    *,
    worker_id: int,
    page_writer: "_InPlacePageWriter",
    preferred_api_key: str | None = None,
) -> tuple[int, str]:
    """Extract one PDF page in a worker thread (opens its own document handle)."""
    import fitz

    actual_page = page_number + 1
    page_label = f"page {actual_page}/{total_pages}"
    key_label = _api_key_label(preferred_api_key)
    _log(f"[Worker {worker_id}] ▶ Started {page_label} ({key_label})")

    doc = fitz.open(stream=file_bytes, filetype="pdf")
    try:
        page = doc.load_page(page_number)
        page_text = _extract_pdf_page_text(page)

        if not page_text:
            _log(f"[Worker {worker_id}] 🖼 {page_label} → Gemini OCR")
            pix = page.get_pixmap(alpha=False)
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                temp_path = Path(tmp.name)
            try:
                pix.save(str(temp_path))
                page_text = process_image(
                    temp_path,
                    preferred_api_key=preferred_api_key,
                    worker_id=worker_id,
                    page_label=page_label,
                ).strip()
            finally:
                if temp_path.exists():
                    temp_path.unlink()
        else:
            _log(f"[Worker {worker_id}] 📄 {page_label} → native text")

        formatted_text = f"--- Page {actual_page} ---\n{page_text or ''}"
        page_writer.add(actual_page, formatted_text)
        return actual_page, formatted_text
    finally:
        doc.close()


def _use_temp_output_dir() -> bool:
    """Use /tmp on cloud hosts (Render, etc.) where the app dir is not writable."""
    return bool(os.getenv("RENDER") or os.getenv("VERCEL"))


OUTPUT_BASE_DIR = Path(tempfile.gettempdir()) if _use_temp_output_dir() else BASE_DIR
EXTRACTED_TEXT_DIR = OUTPUT_BASE_DIR / "extracted_text"


def _pages_marked_in_txt(path: Path) -> set[int]:
    """Return all page numbers that already have a --- Page N --- block in the txt."""
    if not path.exists() or path.stat().st_size == 0:
        return set()
    text = path.read_text(encoding="utf-8", errors="replace")
    return {int(m) for m in _PAGE_HEADER_RE.findall(text)}


def _missing_page_numbers(present: set[int], total_pages: int) -> list[int]:
    return [n for n in range(1, total_pages + 1) if n not in present]


def _parse_page_blocks_from_txt(path: Path) -> dict[int, str]:
    """Split an extracted txt into {page_number: '--- Page N ---\\n...'} blocks."""
    if not path.exists() or path.stat().st_size == 0:
        return {}
    text = path.read_text(encoding="utf-8", errors="replace")
    matches = list(_PAGE_HEADER_RE.finditer(text))
    if not matches:
        return {}

    pages: dict[int, str] = {}
    for index, match in enumerate(matches):
        page_num = int(match.group(1))
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        block = text[start:end].strip()
        if block:
            pages[page_num] = block
    return pages


def _write_ordered_pages(path: Path, pages: dict[int, str]) -> None:
    """Rewrite the txt with all known pages sorted 1, 2, 3, … (correct physical order)."""
    ordered = [pages[page_num] for page_num in sorted(pages)]
    content = "\n\n".join(ordered)
    with _FILE_WRITE_LOCK:
        path.write_text(content, encoding="utf-8")
        with path.open("rb+") as out_file:
            out_file.flush()
            try:
                os.fsync(out_file.fileno())
            except OSError:
                pass


def _build_output_txt_path(original_filename: str) -> Path:
    """Create deterministic output path: <original filename>.txt"""
    source_name = Path(original_filename).name
    txt_name = f"{source_name}.txt"
    EXTRACTED_TEXT_DIR.mkdir(parents=True, exist_ok=True)
    return EXTRACTED_TEXT_DIR / txt_name


class _InPlacePageWriter:
    """Keep pages in memory and rewrite the txt in page order after each new page."""

    def __init__(self, path: Path, *, total_pages: int, initial_pages: dict[int, str]) -> None:
        self.path = path.resolve()
        self.total_pages = total_pages
        self._pages = dict(initial_pages)
        self._lock = threading.Lock()
        if self._pages:
            _write_ordered_pages(self.path, self._pages)
            _log(f"📑 Sorted {len(self._pages)} existing page(s) into correct order")

    def add(self, page_num: int, block: str) -> None:
        with self._lock:
            if page_num in self._pages:
                return
            self._pages[page_num] = block.strip()
            _write_ordered_pages(self.path, self._pages)
            _log(
                f"💾 Saved page {page_num}/{self.total_pages} "
                f"in order ({len(self._pages)} pages in file) → {self.path.name}"
            )

    def flush(self) -> None:
        return


def _write_full_text(path: Path, text: str) -> None:
    path.write_text(text or "", encoding="utf-8")


def _extract_pdf_page_text(page) -> str:
    words = page.get_text("words", sort=True) or []
    if not words:
        return ""

    lines: list[list[tuple[float, float, float, str]]] = []
    current_line: list[tuple[float, float, float, str]] = []
    current_y: float | None = None
    line_tolerance = 4.0

    for word in words:
        x0, y0, x1, _y1, text = word[:5]
        if current_y is None or abs(y0 - current_y) <= line_tolerance:
            current_line.append((float(x0), float(y0), float(x1), str(text)))
            current_y = float(y0) if current_y is None else current_y
            continue

        lines.append(current_line)
        current_line = [(float(x0), float(y0), float(x1), str(text))]
        current_y = float(y0)

    if current_line:
        lines.append(current_line)

    lines.sort(key=lambda line: (min(item[1] for item in line), min(item[0] for item in line)))

    page_lines: list[str] = []
    for line in lines:
        line.sort(key=lambda item: item[0])
        rendered_parts: list[str] = []
        previous_x1: float | None = None

        for x0, _y0, x1, text in line:
            if previous_x1 is not None and x0 - previous_x1 > 28:
                rendered_parts.append("|")
            rendered_parts.append(text)
            previous_x1 = x1

        page_lines.append(" ".join(rendered_parts).strip())

    return "\n".join(line for line in page_lines if line.strip()).strip()


def extract_from_pdf(file_bytes: bytes, original_filename: str) -> tuple[str, int]:

    try:
        import fitz

    except ImportError as exc:

        raise RuntimeError(
            "PyMuPDF is required for PDF extraction. "
            "Install pymupdf."
        ) from exc

    # Open PDF once to get page count (each worker opens its own handle).
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    total_pages = len(doc)
    doc.close()

    output_txt_path = _build_output_txt_path(original_filename)
    existing_pages = _parse_page_blocks_from_txt(output_txt_path)
    pages_on_disk = set(existing_pages)

    if pages_on_disk and max(pages_on_disk) > total_pages:
        output_txt_path.unlink()
        existing_pages = {}
        pages_on_disk = set()

    missing_nums = _missing_page_numbers(pages_on_disk, total_pages)
    if not missing_nums:
        _log(
            f"\n⏩ All {total_pages} page(s) already in "
            f"{output_txt_path.name}; skipping extraction."
        )
        final_text = output_txt_path.read_text(encoding="utf-8").strip()
        return final_text, total_pages

    if output_txt_path.exists() and not pages_on_disk:
        # Leftover file without our page markers (e.g. old format) — avoid appending duplicates.
        output_txt_path.unlink()
        missing_nums = list(range(1, total_pages + 1))

    if pages_on_disk:
        preview = ", ".join(str(n) for n in missing_nums[:12])
        if len(missing_nums) > 12:
            preview += ", …"
        _log(
            f"\n⏩ Resume: {len(pages_on_disk)} page(s) already saved, "
            f"{len(missing_nums)} missing — will extract only: {preview}"
        )

    pending_pages = [page_number for page_number in range(total_pages) if (page_number + 1) in missing_nums]
    page_writer = _InPlacePageWriter(
        output_txt_path,
        total_pages=total_pages,
        initial_pages=existing_pages,
    )

    if pending_pages:
        workers = min(PARALLEL_WORKERS, len(pending_pages))
        _log(
            f"\n⚡ Parallel mode: {workers} worker thread(s), "
            f"{len(API_KEYS)} API key(s), {len(pending_pages)} page(s) to process"
        )
        _log(f"📁 Incremental save (sorted by page) → {output_txt_path.resolve()}")
        _log(f"📌 Pages already in file: {len(existing_pages)}")
        if workers == 1:
            _log("⚠️ Only 1 worker active — set PARALLEL_WORKERS=4 in .env for more speed")

        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [
                executor.submit(
                    _process_pdf_page,
                    file_bytes,
                    page_number,
                    total_pages,
                    worker_id=(idx % workers) + 1,
                    page_writer=page_writer,
                    preferred_api_key=API_KEYS[idx % len(API_KEYS)] if API_KEYS else None,
                )
                for idx, page_number in enumerate(pending_pages)
            ]

            for future in as_completed(futures):
                actual_page, _formatted_text = future.result()
                _log(f"✅ Finished page {actual_page}/{total_pages}")

        page_writer.flush()

    final_text = output_txt_path.read_text(encoding="utf-8").strip()

    return final_text, total_pages


def extract_from_docx(file_bytes: bytes) -> str:
    try:
        from docx import Document
    except ImportError as exc:
        raise RuntimeError("python-docx is required for DOCX extraction.") from exc

    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
        tmp.write(file_bytes)
        temp_path = Path(tmp.name)
    try:
        document = Document(str(temp_path))
        paragraphs = [p.text.strip() for p in document.paragraphs if p.text and p.text.strip()]
        return "\n".join(paragraphs).strip()
    finally:
        if temp_path.exists():
            temp_path.unlink()


def extract_from_doc(file_bytes: bytes) -> str:
    with tempfile.TemporaryDirectory() as temp_dir:
        doc_path = Path(temp_dir) / "input.doc"
        txt_path = Path(temp_dir) / "output.txt"
        doc_path.write_bytes(file_bytes)

        try:
            import win32com.client  # type: ignore
        except ImportError as exc:
            raise RuntimeError("pywin32 is required for .doc extraction on Windows.") from exc

        word = document = None
        try:
            word = win32com.client.DispatchEx("Word.Application")
            word.Visible = False
            word.DisplayAlerts = 0
            document = word.Documents.Open(str(doc_path))
            document.SaveAs(str(txt_path), FileFormat=7)
        finally:
            if document is not None:
                try:
                    document.Close(False)
                except Exception:
                    pass
            if word is not None:
                try:
                    word.Quit()
                except Exception:
                    pass

        for encoding in ("utf-16", "utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                content = txt_path.read_text(encoding=encoding).strip()
                if content:
                    return content
            except UnicodeError:
                continue
        return txt_path.read_text(encoding="utf-8", errors="ignore").strip()


def extract_from_excel(file_bytes: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel extraction.") from exc

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        temp_path = Path(tmp.name)
    try:
        workbook = load_workbook(str(temp_path), data_only=True)
        parts: list[str] = []
        for sheet in workbook.worksheets:
            parts.append(f"--- Sheet: {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    parts.append(" | ".join(cells))
        return "\n".join(parts).strip()
    finally:
        if temp_path.exists():
            temp_path.unlink()


@app.get("/")
async def home() -> FileResponse:
    return FileResponse(str(STATIC_DIR / "index.html"))


@app.get("/styles.css")
async def styles() -> FileResponse:
    return FileResponse(str(STATIC_DIR / "styles.css"))


@app.get("/app.js")
async def app_script() -> FileResponse:
    return FileResponse(str(STATIC_DIR / "app.js"))


@app.post("/api/extract")
async def extract_text(
    file: UploadFile = File(...)
) -> JSONResponse:

    if not file.filename:

        raise HTTPException(
            status_code=400,
            detail="No file uploaded."
        )

    suffix = Path(
        file.filename
    ).suffix.lower()

    content = await file.read()

    if not content:

        raise HTTPException(
            status_code=400,
            detail="Empty file."
        )

    output_txt_path = _build_output_txt_path(file.filename)
    total_pages: int | None = None
    pages_extracted: int | None = None

    try:

        # =========================
        # IMAGE FILES
        # =========================
        if suffix in SUPPORTED_EXTENSIONS:

            with tempfile.NamedTemporaryFile(
                suffix=suffix,
                delete=False
            ) as tmp:

                tmp.write(content)

                temp_path = Path(tmp.name)

            try:

                text = process_image(temp_path)

            finally:

                if temp_path.exists():
                    temp_path.unlink()

            _write_full_text(output_txt_path, text)

        # =========================
        # PDF FILES
        # =========================
        elif suffix in PDF_EXTENSIONS:

            text, total_pages = await asyncio.to_thread(
                extract_from_pdf,
                content,
                file.filename,
            )
            pages_extracted = total_pages

        # =========================
        # DOCX FILES
        # =========================
        elif suffix in WORD_EXTENSIONS:

            text = extract_from_docx(content)
            _write_full_text(output_txt_path, text)

        # =========================
        # DOC FILES
        # =========================
        elif suffix in LEGACY_WORD_EXTENSIONS:

            text = extract_from_doc(content)
            _write_full_text(output_txt_path, text)

        # =========================
        # EXCEL FILES
        # =========================
        elif suffix in EXCEL_EXTENSIONS:

            text = extract_from_excel(content)
            _write_full_text(output_txt_path, text)

        # =========================
        # UNSUPPORTED FILES
        # =========================
        else:

            raise HTTPException(
                status_code=400,
                detail=(
                    "Unsupported file. "
                    "Use image, PDF, DOC/DOCX, "
                    "or XLSX-family formats."
                ),
            )

    except HTTPException:
        raise

    except Exception as exc:

        raise HTTPException(
            status_code=500,
            detail=f"Extraction failed: {exc}"
        ) from exc

    return JSONResponse(
        {
            "filename": file.filename,
            "characters": len(text or ""),
            "text": text or "",
            "pages_extracted": pages_extracted,
            "total_pages": total_pages,
            "structured_rows": [],
            "table_html": [],
        }
    )
@app.post("/api/send-extracted")
async def send_extracted_text(payload: SendExtractedPayload) -> JSONResponse:
    try:
        response = requests.post(
            str(payload.endpoint),
            json={"text": payload.text},
            timeout=30,
        )
        response.raise_for_status()
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"Failed to send extracted text: {exc}") from exc

    response_data: dict | list | str
    try:
        response_data = response.json()
    except ValueError:
        response_data = response.text

    return JSONResponse(
        {
            "endpoint": str(payload.endpoint),
            "status_code": response.status_code,
            "response": response_data,
        }
    )


@app.get("/api/fetch-data")
async def fetch_data(endpoint: HttpUrl) -> JSONResponse:
    try:
        response = requests.get(str(endpoint), timeout=30)
        response.raise_for_status()
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch data: {exc}") from exc

    try:
        data = response.json()
    except ValueError:
        data = {"raw": response.text}

    return JSONResponse(
        {
            "endpoint": str(endpoint),
            "status_code": response.status_code,
            "data": data,
        }
    )


@app.post("/api/process-extracted")
async def process_extracted(payload: ProcessExtractedPayload) -> JSONResponse:
    if not PROCESSOR_POST_ENDPOINT or not PROCESSOR_GET_ENDPOINT:
        raise HTTPException(
            status_code=500,
            detail="Missing PROCESSOR_POST_ENDPOINT or PROCESSOR_GET_ENDPOINT in environment.",
        )

    try:
        post_response = requests.post(
            PROCESSOR_POST_ENDPOINT,
            json={"text": payload.text},
            timeout=30,
        )
        post_response.raise_for_status()
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"Failed sending text to processor: {exc}") from exc

    last_error: Exception | None = None
    for _ in range(12):
        try:
            get_response = requests.get(PROCESSOR_GET_ENDPOINT, timeout=30)
            get_response.raise_for_status()
            try:
                processed_data = get_response.json()
            except ValueError:
                processed_data = {"raw": get_response.text}

            return JSONResponse(
                {
                    "post_endpoint": PROCESSOR_POST_ENDPOINT,
                    "get_endpoint": PROCESSOR_GET_ENDPOINT,
                    "processed_data": processed_data,
                }
            )
        except requests.RequestException as exc:
            last_error = exc
            time.sleep(1)

    raise HTTPException(
        status_code=504,
        detail=f"Processor did not return data in time: {last_error}",
    )


